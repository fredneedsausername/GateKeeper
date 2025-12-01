from flask import Flask, request, session, redirect, render_template, jsonify, flash, send_file
from waitress import serve
from psycopg.rows import dict_row
import os
from psycopg_pool import ConnectionPool
from functools import wraps
from datetime import datetime, timedelta
import base64
from typing import TypeVar, ParamSpec, Callable, Concatenate, Any
import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# Type variables for typing the decorator
P = ParamSpec('P')
R = TypeVar('R')
CursorType = Any

def auth_required(fn):
    @wraps(fn)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect("/login")
        return fn(*args, **kwargs)
    return decorated

def get_env(env_var: str):
    ret = os.getenv(env_var)
    if not ret:
        raise Exception(f"{env_var} not found")
    return ret

flask_env = get_env("FLASK_ENV")

db_url = get_env("DATABASE_URL")
db_pool = ConnectionPool(
    conninfo=db_url,
    min_size=4,
    max_size=20,
    timeout=30.0,
    configure=lambda conn: setattr(conn, 'row_factory', dict_row)
)

def connected_to_database(fn: Callable[Concatenate[CursorType, P], R]) -> Callable[P, R]:
    @wraps(fn)
    def wrapped_function(*args: P.args, **kwargs: P.kwargs) -> R:
        with db_pool.connection() as conn:
            with conn.cursor() as curs:
                try:
                    # Explicitly start transaction
                    conn.execute("BEGIN")
                    ret = fn(curs, *args, **kwargs)
                    # Commit transaction on success
                    conn.commit()
                    return ret
                except Exception:
                    # Rollback transaction on any error
                    conn.rollback()
                    raise
    return wrapped_function

app = Flask(__name__)

@app.template_filter('b64encode')
def b64encode_filter(data):
    """Base64 encode filter for Jinja2"""
    if isinstance(data, str):
        data = data.encode('utf-8')
    return base64.b64encode(data).decode('ascii')

@app.template_filter('datetime_format')
def datetime_format_filter(dt):
    """Format datetime for display"""
    if dt is None:
        return ''
    if hasattr(dt, 'strftime'):
        return dt.strftime('%d/%m/%Y %H:%M:%S')
    return str(dt)

@app.context_processor
def inject_global_vars():
    return {
        'username': session.get("user", "Non riconosciuto"),
    }

secret_key = get_env("SECRET_KEY")
app.secret_key = str(secret_key)

# ===== UTILITY FUNCTIONS =====

def validate_timestamp(timestamp_str, field_name):
    """
    Validate that a timestamp string represents a date between year 1000 and 3000.
    Returns (is_valid, error_message)
    """
    if not timestamp_str:
        return True, None  # Empty timestamps are allowed (at least one must be provided, checked elsewhere)
    
    try:
        dt = datetime.fromisoformat(timestamp_str.replace('T', ' '))
        
        if dt.year < 1000:
            return False, f"{field_name} deve essere dopo l'anno 1000"
        
        if dt.year > 3000:
            return False, f"{field_name} deve essere prima dell'anno 3000"
        
        return True, None
        
    except (ValueError, AttributeError):
        return False, f"{field_name} non è un formato valido"

@connected_to_database
def get_ships_for_dropdown(curs):
    curs.execute("SELECT id, name FROM ship ORDER BY name")
    ships = curs.fetchall()
    return [{"value": ship["id"], "label": ship["name"]} for ship in ships]

@connected_to_database
def get_roles_for_dropdown(curs):
    curs.execute("SELECT id, role_name FROM crew_member_roles ORDER BY role_name")
    roles = curs.fetchall()
    return [{"value": role["id"], "label": role["role_name"]} for role in roles]

@connected_to_database
def get_shipyards_for_dropdown(curs):
    curs.execute("SELECT id, name FROM shipyard ORDER BY name")
    shipyards = curs.fetchall()
    return [{"value": sy["id"], "label": sy["name"]} for sy in shipyards]

def get_filtered_data(table_type, filters, page=1, page_size=50):
    """Get filtered data based on table type and filters"""
    
    @connected_to_database
    def fetch_data(curs):
        if table_type == "crew_member":
            # Only show results if there are actual filters applied (not just empty strings)
            has_filters = any(v for v in filters.values() if v and str(v).strip())
            if not has_filters:
                return [], 0
                
            # Base query parts
            select_fields = """
                cm.id,
                cm.name as crew_member_name,
                cm.phone_number,
                cr.role_name,
                s.name as ship_name,
                t.mac_address as tag_name,
                t.remaining_battery as battery_level
            """
            
            from_where = """
                FROM crew_member cm
                LEFT JOIN crew_member_roles cr ON cm.role_id = cr.id
                LEFT JOIN ship s ON cm.ship_id = s.id
                LEFT JOIN tag t ON cm.tag_id = t.id
                WHERE 1=1
            """
            
            params = []
            filter_conditions = ""
            
            if filters.get('crew_name') and filters['crew_name'].strip():
                filter_conditions += " AND LOWER(cm.name) LIKE LOWER(%s)"
                params.append(f"%{filters['crew_name'].strip()}%")
            
            if filters.get('ship_name') and filters['ship_name'].strip():
                filter_conditions += " AND LOWER(s.name) LIKE LOWER(%s)"
                params.append(f"%{filters['ship_name'].strip()}%")
            
            if filters.get('role_id') and filters['role_id'].strip():
                try:
                    role_id = int(filters['role_id'])
                    filter_conditions += " AND cm.role_id = %s"
                    params.append(role_id)
                except ValueError:
                    pass  # Skip invalid filter
            
            if filters.get('ship_id') and filters['ship_id'].strip():
                try:
                    ship_id = int(filters['ship_id'])
                    filter_conditions += " AND cm.ship_id = %s"
                    params.append(ship_id)
                except ValueError:
                    pass  # Skip invalid filter
            
            # Count total
            count_query = f"SELECT COUNT(*) as count {from_where} {filter_conditions}"
            curs.execute(count_query, params)
            result = curs.fetchone()
            total = result['count'] if result else 0
            
            # Get data with pagination
            data_query = f"SELECT {select_fields} {from_where} {filter_conditions} ORDER BY cm.name LIMIT %s OFFSET %s"
            params.extend([page_size, (page - 1) * page_size])
            
            curs.execute(data_query, params)
            items = curs.fetchall()
            
            return items, total
            
        elif table_type == "ship":
            # Only show results if there is a filter applied (not just empty string)
            has_filters = any(v for v in filters.values() if v and str(v).strip())
            if not has_filters:
                return [], 0

            # Base query parts
            from_where = "FROM ship WHERE 1=1"
            params = []
            filter_conditions = ""
            
            if filters.get('ship_name') and filters['ship_name'].strip():
                filter_conditions += " AND LOWER(name) LIKE LOWER(%s)"
                params.append(f"%{filters['ship_name'].strip()}%")
            
            # Count total
            count_query = f"SELECT COUNT(*) as count {from_where} {filter_conditions}"
            curs.execute(count_query, params)
            result = curs.fetchone()
            total = result['count'] if result else 0
            
            # Get data with pagination
            data_query = f"SELECT id, name {from_where} {filter_conditions} ORDER BY name LIMIT %s OFFSET %s"
            params.extend([page_size, (page - 1) * page_size])
            
            curs.execute(data_query, params)
            items = curs.fetchall()
            
            return items, total
            
        elif table_type == "tag":
            # Base query parts
            select_fields = """
                t.id,
                t.mac_address as name,
                t.remaining_battery,
                cm.name as crew_member_name
            """
            
            from_where = """
                FROM tag t
                LEFT JOIN crew_member cm ON t.id = cm.tag_id
                WHERE 1=1
            """
            
            params = []
            filter_conditions = ""
            
            if filters.get('assigned') and not filters.get('vacant'):
                filter_conditions += " AND cm.id IS NOT NULL"
            elif filters.get('vacant') and not filters.get('assigned'):
                filter_conditions += " AND cm.id IS NULL"
            elif not filters.get('assigned') and not filters.get('vacant'):
                # No checkboxes selected - return empty
                return [], 0
            
            # Filter by tag name if provided
            if filters.get('tag_name') and filters['tag_name'].strip():
                filter_conditions += " AND LOWER(t.mac_address) LIKE LOWER(%s)"
                params.append(f"%{filters['tag_name'].strip()}%")
            
            # Count total
            count_query = f"SELECT COUNT(*) as count {from_where} {filter_conditions}"
            curs.execute(count_query, params)
            result = curs.fetchone()
            total = result['count'] if result else 0
            
            # Get data with pagination
            data_query = f"SELECT {select_fields} {from_where} {filter_conditions} ORDER BY t.remaining_battery ASC LIMIT %s OFFSET %s"
            params.extend([page_size, (page - 1) * page_size])
            
            curs.execute(data_query, params)
            items = curs.fetchall()
            
            return items, total
            
        elif table_type == "unassigned_tag_entry":
            # Base query parts
            select_fields = """
                ute.id,
                s.name as shipyard_name,
                t.mac_address as tag_name,
                t.remaining_battery as battery_level,
                ute.advertisement_timestamp,
                CASE WHEN ute.is_entering THEN 'Ingresso' ELSE 'Uscita' END as entry_type
            """
            
            from_where = """
                FROM unassigned_tag_entry ute
                JOIN tag t ON ute.tag_id = t.id
                JOIN shipyard s ON ute.shipyard_id = s.id
                WHERE 1=1
            """
            
            params = []
            filter_conditions = ""
            
            # Convert string timestamps to datetime objects if needed
            if filters.get('start_timestamp'):
                start_ts = filters['start_timestamp']
                if isinstance(start_ts, str):
                    try:
                        timestamp_str = start_ts.replace('T', ' ')
                        # datetime-local inputs give 'YYYY-MM-DD HH:MM' without seconds
                        # fromisoformat needs 'YYYY-MM-DD HH:MM:SS'
                        if timestamp_str.count(':') == 1:
                            timestamp_str += ':00'
                        start_ts = datetime.fromisoformat(timestamp_str)
                    except:
                        start_ts = datetime.now() - timedelta(hours=24)
                filter_conditions += " AND ute.advertisement_timestamp >= %s"
                params.append(start_ts)
            
            if filters.get('end_timestamp'):
                end_ts = filters['end_timestamp']
                if isinstance(end_ts, str):
                    try:
                        timestamp_str = end_ts.replace('T', ' ')
                        # datetime-local inputs give 'YYYY-MM-DD HH:MM' without seconds
                        # fromisoformat needs 'YYYY-MM-DD HH:MM:SS'
                        if timestamp_str.count(':') == 1:
                            timestamp_str += ':00'
                        end_ts = datetime.fromisoformat(timestamp_str)
                    except:
                        end_ts = datetime.now()
                filter_conditions += " AND ute.advertisement_timestamp <= %s"
                params.append(end_ts)
            
            if filters.get('shipyard_id') and filters['shipyard_id'].strip():
                try:
                    shipyard_id = int(filters['shipyard_id'])
                    filter_conditions += " AND ute.shipyard_id = %s"
                    params.append(shipyard_id)
                except ValueError:
                    pass  # Skip invalid filter
            
            if filters.get('tag_name') and filters['tag_name'].strip():
                filter_conditions += " AND LOWER(t.mac_address) LIKE LOWER(%s)"
                params.append(f"%{filters['tag_name'].strip()}%")
            
            # Count total
            count_query = f"SELECT COUNT(*) as count {from_where} {filter_conditions}"
            curs.execute(count_query, params)
            result = curs.fetchone()
            total = result['count'] if result else 0
            
            # Get data with pagination
            data_query = f"SELECT {select_fields} {from_where} {filter_conditions} ORDER BY ute.advertisement_timestamp DESC LIMIT %s OFFSET %s"
            params.extend([page_size, (page - 1) * page_size])
            
            curs.execute(data_query, params)
            items = curs.fetchall()
            
            return items, total
            
        elif table_type == "permanence_log":
            # Base query parts
            select_fields = """
                pl.id,
                s.name as shipyard_name,
                current_tag.mac_address as current_tag_name,
                current_tag.remaining_battery as current_battery_level,
                ship.name as ship_name,
                cm.name as crew_member_name,
                cr.role_name,
                pl.entry_timestamp,
                pl.leave_timestamp
            """
            
            from_where = """
                FROM permanence_log pl
                JOIN crew_member cm ON pl.crew_member_id = cm.id
                JOIN shipyard s ON pl.shipyard_id = s.id
                LEFT JOIN crew_member_roles cr ON cm.role_id = cr.id
                LEFT JOIN ship ship ON cm.ship_id = ship.id
                LEFT JOIN tag current_tag ON cm.tag_id = current_tag.id
                WHERE 1=1
            """
            
            params = []
            filter_conditions = ""
            
            # Convert string timestamps to datetime objects and apply time filters
            if filters.get('start_timestamp') and filters.get('end_timestamp'):
                start_ts = filters['start_timestamp']
                end_ts = filters['end_timestamp']
                
                if isinstance(start_ts, str):
                    try:
                        timestamp_str = start_ts.replace('T', ' ')
                        # datetime-local inputs give 'YYYY-MM-DD HH:MM' without seconds
                        # fromisoformat needs 'YYYY-MM-DD HH:MM:SS'
                        if timestamp_str.count(':') == 1:
                            timestamp_str += ':00'
                        start_ts = datetime.fromisoformat(timestamp_str)
                    except:
                        start_ts = datetime.now() - timedelta(hours=24)
                        
                if isinstance(end_ts, str):
                    try:
                        timestamp_str = end_ts.replace('T', ' ')
                        # datetime-local inputs give 'YYYY-MM-DD HH:MM' without seconds
                        # fromisoformat needs 'YYYY-MM-DD HH:MM:SS'
                        if timestamp_str.count(':') == 1:
                            timestamp_str += ':00'
                        end_ts = datetime.fromisoformat(timestamp_str)
                    except:
                        end_ts = datetime.now()
                
                # Show log if entry OR exit is within the date range
                filter_conditions += """
                    AND (
                        (pl.entry_timestamp >= %s AND pl.entry_timestamp <= %s)
                        OR (pl.leave_timestamp >= %s AND pl.leave_timestamp <= %s)
                    )
                """
                params.extend([start_ts, end_ts, start_ts, end_ts])
            
            if filters.get('shipyard_id') and filters['shipyard_id'].strip():
                try:
                    shipyard_id = int(filters['shipyard_id'])
                    filter_conditions += " AND pl.shipyard_id = %s"
                    params.append(shipyard_id)
                except ValueError:
                    pass  # Skip invalid filter
            
            if filters.get('ship_id') and filters['ship_id'].strip():
                try:
                    ship_id = int(filters['ship_id'])
                    filter_conditions += " AND cm.ship_id = %s"
                    params.append(ship_id)
                except ValueError:
                    pass  # Skip invalid filter
            
            if filters.get('crew_name') and filters['crew_name'].strip():
                filter_conditions += " AND LOWER(cm.name) LIKE LOWER(%s)"
                params.append(f"%{filters['crew_name'].strip()}%")
            
            # Count total
            count_query = f"SELECT COUNT(*) as count {from_where} {filter_conditions}"
            try:
                curs.execute(count_query, params)
                result = curs.fetchone()
                total = result['count'] if result else 0
            except Exception as e:
                print(f"Error in count query: {e}")
                print(f"Query: {count_query}")
                print(f"Params: {params}")
                total = 0
            
            # Get data with pagination
            data_query = f"SELECT {select_fields} {from_where} {filter_conditions} ORDER BY cm.name ASC LIMIT %s OFFSET %s"
            params_with_pagination = params + [page_size, (page - 1) * page_size]
            
            try:
                curs.execute(data_query, params_with_pagination)
                items = curs.fetchall()
            except Exception as e:
                print(f"Error in data query: {e}")
                print(f"Query: {data_query}")
                print(f"Params: {params_with_pagination}")
                items = []
            
            return items, total
            
        return [], 0
    
    return fetch_data()

# ===== HTMX TABLE CONFIG HELPER =====

def create_table_config(table_type, filters, page, request_path):
    """Helper function to create table configuration for both full and partial requests"""
    data, total_count = get_filtered_data(table_type, filters, page)
    
    if table_type == "crew_member":
        return {
            "title": "Gestione Crew",
            "description": "Visualizza e gestisci i membri dell'equipaggio",
            "columns": [
                {"key": "tag_name", "label": "Tag", "type": "text"},
                {"key": "battery_level", "label": "🔋%", "type": "battery"},
                {"key": "ship_name", "label": "Barca", "type": "text"},
                {"key": "crew_member_name", "label": "Equipaggio", "type": "text"},
                {"key": "role_name", "label": "Ruolo", "type": "text"},
                {"key": "phone_number", "label": "Telefono", "type": "text"}
            ],
            "text_filters": [
                {"key": "crew_name", "label": "Nominativo", "placeholder": "Cerca per nome..."}
            ],
            "searchable_select_filters": [
                {"key": "ship_id", "label": "Barca", "placeholder": "Cerca barca...", "search_endpoint": "/api/ships/filter"},
                {"key": "role_id", "label": "Ruolo", "placeholder": "Cerca ruolo...", "search_endpoint": "/api/roles/filter"}
            ],
            "allow_add": True,
            "allow_edit": True,
            "allow_delete": True,
            "add_button_text": "Aggiungi Crew",
            "empty_message": "Applica dei filtri per visualizzare i membri dell'equipaggio.",
            "add_url": "/crew/add",
            "edit_url": "/crew/edit/{id}",
            "delete_url": "/api/crew/delete/{id}",
            "data": data,
            "total_count": total_count,
            "page": page
        }
    
    elif table_type == "ship":
        return {
            "title": "Gestione Barche",
            "description": "Visualizza e gestisci le barche",
            "columns": [
                {"key": "name", "label": "Barca", "type": "text"}
            ],
            "text_filters": [
                {"key": "ship_name", "label": "Nome Barca", "placeholder": "Cerca per nome..."}
            ],
            "allow_add": True,
            "allow_edit": True,
            "allow_delete": True,
            "add_button_text": "Aggiungi Barca",
            "empty_message": "Applica dei filtri per visualizzare le barche.",
            "add_url": "/navi/add",
            "edit_url": "/navi/edit/{id}",
            "delete_url": "/api/ships/delete/{id}",
            "data": data,
            "total_count": total_count,
            "page": page
        }
    
    elif table_type == "tag":
        return {
            "title": "Gestione Tag",
            "description": "Visualizza i tag",
            "columns": [
                {"key": "name", "label": "Tag", "type": "text"},
                {"key": "remaining_battery", "label": "🔋%", "type": "battery"},
                {"key": "crew_member_name", "label": "Equipaggio", "type": "text"}
            ],
            "text_filters": [
                {"key": "tag_name", "label": "Nome Tag", "placeholder": "Cerca per nome..."}
            ],
            "checkbox_filters": [
                {"key": "assigned", "label": "Assegnati"},
                {"key": "vacant", "label": "Vacanti"}
            ],
            "allow_add": False,
            "allow_edit": False,
            "allow_delete": False,
            "add_button_text": "Aggiungi Tag",
            "empty_message": "Seleziona almeno un filtro per visualizzare i tag.",
            "add_url": "/tag/add",
            "edit_url": "/tag/edit/{id}",
            "delete_url": "/api/tags/delete/{id}",
            "data": data,
            "total_count": total_count,
            "page": page
        }
    
    elif table_type == "unassigned_tag_entry":
        return {
            "title": "Entry dei Tag",
            "description": "Visualizza gli eventi dei tag non assegnati",
            "columns": [
                {"key": "shipyard_name", "label": "Cantiere", "type": "text"},
                {"key": "tag_name", "label": "Tag", "type": "text"},
                {"key": "battery_level", "label": "🔋%", "type": "battery"},
                {"key": "advertisement_timestamp", "label": "Passaggio", "type": "datetime"},
                {"key": "entry_type", "label": "Tipologia", "type": "text"}
            ],
            "date_filters": [
                {
                    "key": "start_timestamp", 
                    "label": "Data Inizio",
                    "default_value": (datetime.now() - timedelta(hours=24)).strftime('%Y-%m-%dT%H:%M')
                },
                {
                    "key": "end_timestamp", 
                    "label": "Data Fine",
                    "default_value": datetime.now().strftime('%Y-%m-%dT%H:%M')
                }
            ],
            "searchable_select_filters": [
                {"key": "shipyard_id", "label": "Cantiere", "placeholder": "Cerca cantiere...", "search_endpoint": "/api/shipyards/filter"}
            ],
            "text_filters": [
                {"key": "tag_name", "label": "Nome Tag", "placeholder": "Cerca tag esatto..."}
            ],
            "allow_add": False,
            "allow_edit": False,
            "allow_delete": True,
            "allow_export": True,  # ADD THIS LINE
            "empty_message": "Nessuna entry trovata nel periodo selezionato.",
            "delete_url": "/api/entries/delete/{id}",
            "export_url": "/entry/export",  # ADD THIS LINE
            "data": data,
            "total_count": total_count,
            "page": page
        }

    elif table_type == "permanence_log":
        return {
            "title": "Log Permanenze",
            "description": "Visualizza i log di permanenza nei cantieri",
            "columns": [
                {"key": "shipyard_name", "label": "Cantiere", "type": "text"},
                {"key": "current_tag_name", "label": "Tag", "type": "text"},
                {"key": "current_battery_level", "label": "🔋%", "type": "battery"},
                {"key": "ship_name", "label": "Barca", "type": "text"},
                {"key": "crew_member_name", "label": "Equipaggio", "type": "text"},
                {"key": "role_name", "label": "Ruolo", "type": "text"},
                {"key": "entry_timestamp", "label": "Entrata", "type": "datetime"},
                {"key": "leave_timestamp", "label": "Uscita", "type": "datetime"}
            ],
            "date_filters": [
                {
                    "key": "start_timestamp", 
                    "label": "Data Inizio",
                    "default_value": (datetime.now() - timedelta(hours=24)).strftime('%Y-%m-%dT%H:%M')
                },
                {
                    "key": "end_timestamp", 
                    "label": "Data Fine",
                    "default_value": datetime.now().strftime('%Y-%m-%dT%H:%M')
                }
            ],
            "searchable_select_filters": [
                {"key": "shipyard_id", "label": "Cantiere", "placeholder": "Cerca cantiere...", "search_endpoint": "/api/shipyards/filter"},
                {"key": "ship_id", "label": "Barca", "placeholder": "Cerca barca...", "search_endpoint": "/api/ships/filter"}
            ],
            "text_filters": [
                {"key": "crew_name", "label": "Nominativo", "placeholder": "Cerca per nome..."}
            ],
            "allow_add": True,
            "allow_edit": True,
            "allow_delete": True,
            "allow_export": True,
            "add_button_text": "Aggiungi Log",
            "empty_message": "Nessun log trovato nel periodo selezionato.",
            "add_url": "/log/add",
            "edit_url": "/log/edit/{id}",
            "delete_url": "/api/logs/delete/{id}",
            "export_url": "/log/export",
            "data": data,
            "total_count": total_count,
            "page": page
        }

    return {
        "title": "Unknown Table",
        "description": "",
        "columns": [],
        "data": data,
        "total_count": total_count,
        "page": page
    }

# ===== MAIN ROUTES =====

@app.route("/")
@auth_required
def index():
    return redirect("/log")

@app.route("/login", methods=["GET"])
def login_get():
    return render_template("login.html")

@app.route("/login", methods=["POST"])
@connected_to_database
def login_post(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    username = data.get('username')
    password = data.get('password')
    curs.execute("SELECT username FROM users WHERE username=%s AND password=%s", (username, password))
    user = curs.fetchone()
    if user:
        session["user"] = username
        return jsonify({
            "message": "Login effettuato con successo",
            "redirect": "/"
        }), 200
    else:
        return jsonify({
            "error": "Credenziali incorrette"
        }), 401

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ===== CREW MEMBERS ROUTES WITH HTMX =====

@app.route('/crew')
@app.route('/crew/partial')
@auth_required
def crew_page():
    # Get filters from URL parameters
    filters = {
        'crew_name': request.args.get('crew_name', ''),
        'ship_name': request.args.get('ship_name', ''),
        'role_id': request.args.get('role_id', ''),
        'ship_id': request.args.get('ship_id', '')
    }
    
    page = int(request.args.get('page', 1))
    
    try:
        table_config = create_table_config("crew_member", filters, page, request.path)
    except Exception as e:
        print(f"Error in crew_page: {e}")
        table_config = create_table_config("crew_member", {}, 1, request.path)
    
    # Return partial template for HTMX requests
    if request.path.endswith('/partial'):
        return render_template('table_content_partial.html', table_config=table_config)
    
    # Return full page for regular requests
    return render_template('crew.html', table_config=table_config)

@app.route('/crew/add', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def add_crew(curs):
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            ship_id = request.form.get('ship_id') or None
            role_id = request.form.get('role_id') or None
            tag_id = request.form.get('tag_id') or None
            phone_number = request.form.get('phone_number') or None
            if not name:
                flash('Nome è obbligatorio', 'error')
                return redirect(request.url)
            
            # Use database constraint to enforce tag uniqueness
            try:
                curs.execute(
                    "INSERT INTO crew_member (name, ship_id, role_id, tag_id, phone_number) VALUES (%s, %s, %s, %s, %s)",
                    [name, ship_id, role_id, tag_id, phone_number]
                )
                flash('Crew member aggiunto con successo', 'success')
                return redirect('/crew')
            except psycopg.IntegrityError:
                flash('Tag già assegnato a un altro membro dell\'equipaggio', 'error')
                return redirect(request.url)
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiunta: {str(e)}', 'error')
            return redirect(request.url)
    return render_template('crew_add.html')

@app.route('/crew/edit/<int:crew_id>', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def edit_crew(curs, crew_id):
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            ship_id = request.form.get('ship_id') or None
            role_id = request.form.get('role_id') or None
            tag_id = request.form.get('tag_id') or None
            phone_number = request.form.get('phone_number') or None
            if not name:
                flash('Nome è obbligatorio', 'error')
                return redirect(request.url)
            
            # Use database constraint to enforce tag uniqueness
            try:
                curs.execute(
                    "UPDATE crew_member SET name = %s, ship_id = %s, role_id = %s, tag_id = %s, phone_number = %s WHERE id = %s",
                    [name, ship_id, role_id, tag_id, phone_number, crew_id]
                )
                flash('Crew member aggiornato con successo', 'success')
                return redirect('/crew')
            except psycopg.IntegrityError:
                flash('Tag già assegnato a un altro membro dell\'equipaggio', 'error')
                return redirect(request.url)
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
            return redirect(request.url)
    # GET request
    curs.execute(
        """SELECT cm.*, s.name as ship_name, cr.role_name, t.mac_address as tag_name
           FROM crew_member cm
           LEFT JOIN ship s ON cm.ship_id = s.id
           LEFT JOIN crew_member_roles cr ON cm.role_id = cr.id
           LEFT JOIN tag t ON cm.tag_id = t.id
           WHERE cm.id = %s""",
        [crew_id]
    )
    crew_member = curs.fetchone()
    if not crew_member:
        flash('Crew member non trovato', 'error')
        return redirect('/crew')
    return render_template('crew_edit.html', crew_member=crew_member)

@app.route('/api/crew/delete/<int:crew_id>', methods=['DELETE'])
@auth_required
@connected_to_database
def delete_crew(curs, crew_id):
    try:
        curs.execute("DELETE FROM crew_member WHERE id = %s", [crew_id])
        return jsonify({"success": True, "message": "Crew member eliminato con successo"})
    except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
        return jsonify({"success": False, "error": f"Errore database: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ===== SHIPS ROUTES WITH HTMX =====

@app.route('/navi')
@app.route('/navi/partial')
@auth_required
def ships_page():
    # Get filters from URL parameters
    filters = {
        'ship_name': request.args.get('ship_name', '')
    }
    
    page = int(request.args.get('page', 1))
    
    try:
        table_config = create_table_config("ship", filters, page, request.path)
    except Exception as e:
        print(f"Error in ships_page: {e}")
        table_config = create_table_config("ship", {}, 1, request.path)
    
    # Return partial template for HTMX requests
    if request.path.endswith('/partial'):
        return render_template('table_content_partial.html', table_config=table_config)
    
    # Return full page for regular requests
    return render_template('ships.html', table_config=table_config)

@app.route('/navi/add', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def add_ship(curs):
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            if not name:
                flash('Nome è obbligatorio', 'error')
                return redirect(request.url)
            curs.execute("INSERT INTO ship (name) VALUES (%s)", [name])
            flash('Barca aggiunta con successo', 'success')
            return redirect('/navi')
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiunta: {str(e)}', 'error')
            return redirect(request.url)
    return render_template('ship_add.html')

@app.route('/navi/edit/<int:ship_id>', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def edit_ship(curs, ship_id):
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            if not name:
                flash('Nome è obbligatorio', 'error')
                return redirect(request.url)
            curs.execute("UPDATE ship SET name = %s WHERE id = %s", [name, ship_id])
            flash('Barca aggiornata con successo', 'success')
            return redirect('/navi')
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
            return redirect(request.url)
    # GET request
    curs.execute("SELECT * FROM ship WHERE id = %s", [ship_id])
    ship = curs.fetchone()
    if not ship:
        flash('Barca non trovata', 'error')
        return redirect('/navi')
    return render_template('ship_edit.html', ship=ship)

@app.route('/api/ships/delete/<int:ship_id>', methods=['DELETE'])
@auth_required
@connected_to_database
def delete_ship(curs, ship_id):
    try:
        curs.execute("DELETE FROM ship WHERE id = %s", [ship_id])
        return jsonify({"success": True, "message": "Barca eliminata con successo"})
    except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
        return jsonify({"success": False, "error": f"Errore database: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ===== TAGS ROUTES WITH HTMX =====

@app.route('/tag')
@app.route('/tag/partial')
@auth_required
def tags_page():
    # Default both Vacanti and Assegnati when no checkboxes selected
    if not request.args.get('assigned') and not request.args.get('vacant') and not request.path.endswith('/partial'):
        return redirect(f"{request.path}?assigned=1&vacant=1")
    # Get filters from URL parameters
    filters = {
        'assigned': request.args.get('assigned'),
        'vacant': request.args.get('vacant'),
        'tag_name': request.args.get('tag_name', '').strip()
    }
    
    page = int(request.args.get('page', 1))
    
    try:
        table_config = create_table_config("tag", filters, page, request.path)
    except Exception as e:
        print(f"Error in tags_page: {e}")
        table_config = create_table_config("tag", {}, 1, request.path)
    
    # Return partial template for HTMX requests
    if request.path.endswith('/partial'):
        return render_template('table_content_partial.html', table_config=table_config)
    
    # Return full page for regular requests
    return render_template('tags.html', table_config=table_config)

# ===== ENTRIES ROUTES WITH HTMX =====

@app.route('/entry')
@app.route('/entry/partial')
@auth_required
def entries_page():
    # Default to last 24 hours if no dates specified
    now = datetime.now()
    yesterday = now - timedelta(hours=24)
    
    # Get filters from URL parameters with proper defaults
    filters = {
        'start_timestamp': request.args.get('start_timestamp', yesterday.strftime('%Y-%m-%dT%H:%M')),
        'end_timestamp': request.args.get('end_timestamp', now.strftime('%Y-%m-%dT%H:%M')),
        'shipyard_id': request.args.get('shipyard_id', ''),
        'tag_name': request.args.get('tag_name', '')
    }
    
    page = int(request.args.get('page', 1))
    
    try:
        table_config = create_table_config("unassigned_tag_entry", filters, page, request.path)
    except Exception as e:
        print(f"Error in entries_page: {e}")
        table_config = create_table_config("unassigned_tag_entry", {}, 1, request.path)
    
    # Return partial template for HTMX requests
    if request.path.endswith('/partial'):
        return render_template('table_content_partial.html', table_config=table_config)
    
    # Return full page for regular requests
    return render_template('entries.html', table_config=table_config)

@app.route('/api/entries/delete/<int:entry_id>', methods=['DELETE'])
@auth_required
@connected_to_database
def delete_entry(curs, entry_id):
    try:
        curs.execute("DELETE FROM unassigned_tag_entry WHERE id = %s", [entry_id])
        return jsonify({"success": True, "message": "Entry eliminata con successo"})
    except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
        return jsonify({"success": False, "error": f"Errore database: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ===== LOGS ROUTES WITH HTMX =====

@app.route('/log')
@app.route('/log/partial')
@auth_required
def logs_page():
    # Default to last 24 hours if no dates specified
    now = datetime.now()
    yesterday = now - timedelta(hours=24)
    
    # Get filters from URL parameters with proper defaults
    filters = {
        'start_timestamp': request.args.get('start_timestamp', yesterday.strftime('%Y-%m-%dT%H:%M')),
        'end_timestamp': request.args.get('end_timestamp', now.strftime('%Y-%m-%dT%H:%M')),
        'shipyard_id': request.args.get('shipyard_id', ''),
        'ship_id': request.args.get('ship_id', ''),
        'crew_name': request.args.get('crew_name', '')
    }
    
    page = int(request.args.get('page', 1))
    
    try:
        table_config = create_table_config("permanence_log", filters, page, request.path)
    except Exception as e:
        print(f"Error in logs_page: {e}")
        table_config = create_table_config("permanence_log", {}, 1, request.path)
    
    # Return partial template for HTMX requests
    if request.path.endswith('/partial'):
        return render_template('table_content_partial.html', table_config=table_config)
    
    # Return full page for regular requests
    return render_template('logs.html', table_config=table_config)


@app.route('/log/export')
@auth_required
def export_logs():
    """Export filtered logs to Excel file"""
    try:
        # Get filters from URL parameters (no defaults - use what's in URL)
        filters = {
            'start_timestamp': request.args.get('start_timestamp'),
            'end_timestamp': request.args.get('end_timestamp'),
            'shipyard_id': request.args.get('shipyard_id', ''),
            'ship_id': request.args.get('ship_id', ''),
            'crew_name': request.args.get('crew_name', '')
        }
        
        # Fetch up to 10,000 rows matching the filters
        data, total_count = get_filtered_data('permanence_log', filters, page=1, page_size=10000)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Permanenze"
        
        # Define headers (exact order from the table)
        headers = ['Cantiere', 'Tag', '🔋%', 'Barca', 'Equipaggio', 'Ruolo', 'Entrata', 'Uscita']
        
        # Define styles
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Border style
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Cell alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows with minimal styling
        for row_num, row_data in enumerate(data, 2):
            # Cantiere
            cell = ws.cell(row=row_num, column=1, value=row_data.get('shipyard_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Tag
            cell = ws.cell(row=row_num, column=2, value=row_data.get('current_tag_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Battery percentage
            battery = row_data.get('current_battery_level')
            cell = ws.cell(row=row_num, column=3, value=battery if battery is not None else '')
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Barca
            cell = ws.cell(row=row_num, column=4, value=row_data.get('ship_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Equipaggio
            cell = ws.cell(row=row_num, column=5, value=row_data.get('crew_member_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            cell.font = Font(bold=True)
            
            # Ruolo
            cell = ws.cell(row=row_num, column=6, value=row_data.get('role_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Entrata
            entry_ts = row_data.get('entry_timestamp')
            cell = ws.cell(row=row_num, column=7, value=entry_ts.strftime('%d/%m/%Y %H:%M:%S') if entry_ts else '')
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Uscita
            leave_ts = row_data.get('leave_timestamp')
            cell = ws.cell(row=row_num, column=8, value=leave_ts.strftime('%d/%m/%Y %H:%M:%S') if leave_ts else '')
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Auto-adjust column widths with better sizing
        column_widths = {
            'A': 15,  # Cantiere
            'B': 15,  # Tag
            'C': 8,   # Battery
            'D': 15,  # Barca
            'E': 20,  # Equipaggio
            'F': 15,  # Ruolo
            'G': 20,  # Entrata
            'H': 20   # Uscita
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Permanenze.xlsx'
        )
        
    except Exception as e:
        print(f"Error in export_logs: {e}")
        flash(f'Errore durante l\'esportazione: {str(e)}', 'error')
        return redirect('/log')


@app.route('/entry/export')
@auth_required
def export_entries():
    """Export filtered entries to Excel file"""
    try:
        # Get filters from URL parameters (no defaults - use what's in URL)
        filters = {
            'start_timestamp': request.args.get('start_timestamp'),
            'end_timestamp': request.args.get('end_timestamp'),
            'shipyard_id': request.args.get('shipyard_id', ''),
            'tag_name': request.args.get('tag_name', '')
        }
        
        # Fetch up to 10,000 rows matching the filters
        data, total_count = get_filtered_data('unassigned_tag_entry', filters, page=1, page_size=10000)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tag Entries"
        
        # Define headers (exact order from the table)
        headers = ['Cantiere', 'Tag', '🔋%', 'Passaggio', 'Tipologia']
        
        # Define styles
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Border style
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Cell alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows with minimal styling
        for row_num, row_data in enumerate(data, 2):
            # Cantiere
            cell = ws.cell(row=row_num, column=1, value=row_data.get('shipyard_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Tag
            cell = ws.cell(row=row_num, column=2, value=row_data.get('tag_name', ''))
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Battery percentage
            battery = row_data.get('battery_level')
            cell = ws.cell(row=row_num, column=3, value=battery if battery is not None else '')
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Passaggio (advertisement_timestamp)
            adv_ts = row_data.get('advertisement_timestamp')
            cell = ws.cell(row=row_num, column=4, value=adv_ts.strftime('%d/%m/%Y %H:%M:%S') if adv_ts else '')
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Tipologia (entry_type: "Ingresso" or "Uscita")
            cell = ws.cell(row=row_num, column=5, value=row_data.get('entry_type', ''))
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Auto-adjust column widths
        column_widths = {
            'A': 15,  # Cantiere
            'B': 15,  # Tag
            'C': 8,   # Battery
            'D': 20,  # Passaggio
            'E': 12   # Tipologia
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Tag_Entries.xlsx'
        )
        
    except Exception as e:
        print(f"Error in export_entries: {e}")
        flash(f'Errore durante l\'esportazione: {str(e)}', 'error')
        return redirect('/entry')


@app.route('/log/add', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def add_log(curs):
    if request.method == 'POST':
        try:
            crew_member_id = request.form.get('crew_member_id')
            shipyard_id = request.form.get('shipyard_id')
            entry_timestamp = request.form.get('entry_timestamp') or None
            leave_timestamp = request.form.get('leave_timestamp') or None
            
            if not crew_member_id or not shipyard_id:
                flash('Crew member e cantiere sono obbligatori', 'error')
                return redirect(request.url)
            
            if not entry_timestamp and not leave_timestamp:
                flash('Almeno una data (entrata o uscita) è obbligatoria', 'error')
                return redirect(request.url)
            
            # Validate entry timestamp
            is_valid, error_msg = validate_timestamp(entry_timestamp, "Data entrata")
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(request.url)
            
            # Validate leave timestamp
            is_valid, error_msg = validate_timestamp(leave_timestamp, "Data uscita")
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(request.url)
            
            curs.execute(
                "INSERT INTO permanence_log (crew_member_id, shipyard_id, entry_timestamp, leave_timestamp) VALUES (%s, %s, %s, %s)",
                [crew_member_id, shipyard_id, entry_timestamp, leave_timestamp]
            )
            flash('Log aggiunto con successo', 'success')
            return redirect('/log')
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiunta: {str(e)}', 'error')
            return redirect(request.url)
    return render_template('log_add.html')

@app.route('/log/edit/<int:log_id>', methods=['GET', 'POST'])
@auth_required
@connected_to_database
def edit_log(curs, log_id):
    if request.method == 'POST':
        try:
            crew_member_id = request.form.get('crew_member_id')
            entry_timestamp = request.form.get('entry_timestamp') or None
            leave_timestamp = request.form.get('leave_timestamp') or None
            
            if not crew_member_id:
                flash('Crew member è obbligatorio', 'error')
                return redirect(request.url)
            
            if not entry_timestamp and not leave_timestamp:
                flash('Almeno una data (entrata o uscita) è obbligatoria', 'error')
                return redirect(request.url)
            
            # Validate entry timestamp
            is_valid, error_msg = validate_timestamp(entry_timestamp, "Data entrata")
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(request.url)
            
            # Validate leave timestamp
            is_valid, error_msg = validate_timestamp(leave_timestamp, "Data uscita")
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(request.url)
            
            curs.execute(
                "UPDATE permanence_log SET crew_member_id = %s, entry_timestamp = %s, leave_timestamp = %s WHERE id = %s",
                [crew_member_id, entry_timestamp, leave_timestamp, log_id]
            )
            flash('Log aggiornato con successo', 'success')
            return redirect('/log')
        except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
            flash(f'Errore database: {str(e)}', 'error')
            return redirect(request.url)
        except Exception as e:
            flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
            return redirect(request.url)
    # GET request
    curs.execute(
        """SELECT pl.*, cm.name as crew_member_name, s.name as shipyard_name
           FROM permanence_log pl
           JOIN crew_member cm ON pl.crew_member_id = cm.id
           JOIN shipyard s ON pl.shipyard_id = s.id
           WHERE pl.id = %s""",
        [log_id]
    )
    log = curs.fetchone()
    if not log:
        flash('Log non trovato', 'error')
        return redirect('/log')
    return render_template('log_edit.html', log=log)

@app.route('/api/logs/delete/<int:log_id>', methods=['DELETE'])
@auth_required
@connected_to_database
def delete_log(curs, log_id):
    try:
        curs.execute("DELETE FROM permanence_log WHERE id = %s", [log_id])
        return jsonify({"success": True, "message": "Log eliminato con successo"})
    except (psycopg.DatabaseError, psycopg.InterfaceError) as e:
        return jsonify({"success": False, "error": f"Errore database: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ===== SEARCH ENDPOINTS (keep for add/edit forms) =====

@app.route('/api/ships/search', methods=['POST'])
@auth_required
@connected_to_database
def search_ships(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, name FROM ship WHERE LOWER(name) LIKE LOWER(%s) ORDER BY name LIMIT 10",
        [f"%{query}%"]
    )
    ships = curs.fetchall()
    
    return jsonify({"ships": ships})

@app.route('/api/roles/search', methods=['POST'])
@auth_required
@connected_to_database
def search_roles(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, role_name FROM crew_member_roles WHERE LOWER(role_name) LIKE LOWER(%s) ORDER BY role_name LIMIT 10",
        [f"%{query}%"]
    )
    roles = curs.fetchall()
    
    return jsonify({"roles": roles})

@app.route('/api/tags/search', methods=['POST'])
@auth_required
@connected_to_database
def search_tags(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    current_tag_id = data.get('current_tag_id') or None

    # Search only unassigned tags, but include the current tag if editing
    params = [f"%{query}%"]
    subquery = "SELECT tag_id FROM crew_member WHERE tag_id IS NOT NULL"
    if current_tag_id:
        sql = (
            "SELECT t.id, t.mac_address as name "
            "FROM tag t "
            "WHERE LOWER(t.mac_address) LIKE LOWER(%s) "
            "  AND (t.id = %s OR t.id NOT IN (" + subquery + ")) "
            "ORDER BY t.mac_address LIMIT 10"
        )
        params.append(current_tag_id)
    else:
        sql = (
            "SELECT t.id, t.mac_address as name "
            "FROM tag t "
            "WHERE LOWER(t.mac_address) LIKE LOWER(%s) "
            "  AND t.id NOT IN (" + subquery + ") "
            "ORDER BY t.mac_address LIMIT 10"
        )
    curs.execute(sql, params)
    tags = curs.fetchall()
    return jsonify({"tags": tags})

@app.route('/api/crew/search', methods=['POST'])
@auth_required
@connected_to_database
def search_crew(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, name FROM crew_member WHERE LOWER(name) LIKE LOWER(%s) ORDER BY name LIMIT 10",
        [f"%{query}%"]
    )
    crew_members = curs.fetchall()
    
    return jsonify({"crew_members": crew_members})

@app.route('/api/shipyards/search', methods=['POST'])
@auth_required
@connected_to_database
def search_shipyards(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, name FROM shipyard WHERE LOWER(name) LIKE LOWER(%s) ORDER BY name LIMIT 10",
        [f"%{query}%"]
    )
    shipyards = curs.fetchall()
    
    return jsonify({"shipyards": shipyards})

# ===== NEW FILTER SEARCH ENDPOINTS =====

@app.route('/api/ships/filter', methods=['POST'])
@auth_required
@connected_to_database
def filter_ships(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    selected_id = data.get('selected_id')
    
    # If we have a selected_id but no query, return just that item
    if selected_id and not query:
        curs.execute("SELECT id, name FROM ship WHERE id = %s", [selected_id])
        ship = curs.fetchone()
        if ship:
            return jsonify({"options": [{"value": ship["id"], "label": ship["name"]}]})
        return jsonify({"options": []})
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, name FROM ship WHERE LOWER(name) LIKE LOWER(%s) ORDER BY name LIMIT 10",
        [f"%{query}%"]
    )
    ships = curs.fetchall()
    
    options = [{"value": ship["id"], "label": ship["name"]} for ship in ships]
    return jsonify({"options": options})

@app.route('/api/roles/filter', methods=['POST'])
@auth_required
@connected_to_database
def filter_roles(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    selected_id = data.get('selected_id')
    
    # If we have a selected_id but no query, return just that item
    if selected_id and not query:
        curs.execute("SELECT id, role_name FROM crew_member_roles WHERE id = %s", [selected_id])
        role = curs.fetchone()
        if role:
            return jsonify({"options": [{"value": role["id"], "label": role["role_name"]}]})
        return jsonify({"options": []})
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, role_name FROM crew_member_roles WHERE LOWER(role_name) LIKE LOWER(%s) ORDER BY role_name LIMIT 10",
        [f"%{query}%"]
    )
    roles = curs.fetchall()
    
    options = [{"value": role["id"], "label": role["role_name"]} for role in roles]
    return jsonify({"options": options})

@app.route('/api/shipyards/filter', methods=['POST'])
@auth_required
@connected_to_database
def filter_shipyards(curs):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400
        
    query = data.get('query', '').strip()
    selected_id = data.get('selected_id')
    
    # If we have a selected_id but no query, return just that item
    if selected_id and not query:
        curs.execute("SELECT id, name FROM shipyard WHERE id = %s", [selected_id])
        shipyard = curs.fetchone()
        if shipyard:
            return jsonify({"options": [{"value": shipyard["id"], "label": shipyard["name"]}]})
        return jsonify({"options": []})
    
    # Remove the 2-character minimum requirement
    curs.execute(
        "SELECT id, name FROM shipyard WHERE LOWER(name) LIKE LOWER(%s) ORDER BY name LIMIT 10",
        [f"%{query}%"]
    )
    shipyards = curs.fetchall()
    
    options = [{"value": shipyard["id"], "label": shipyard["name"]} for shipyard in shipyards]
    return jsonify({"options": options})


if flask_env == 'development':

    print("mac         |echo |rssi (dbm)|count|scan time   |HH:MM:SS.MMMMMM")
    print("------------+-----+----------+-----+------------+---------------")

    def print_formatted_data(mac, echo, rssi, count, scan_time):

        now = datetime.now()

        # Format as HH:MM:SS.MMMMMM
        time_str = now.strftime("%H:%M:%S.%f")

        print(f"{mac}|{echo.ljust(5)}|{rssi.ljust(10)}|{count.ljust(5)}|{scan_time.ljust(12)}|{time_str.ljust(15)}")

    @app.route('/gateway-endpoint', methods=['POST'])
    def gateway_endpoint():
        json_data = request.json

        data = json_data.get("data")
        if not data:
            return "Invalid gateway message", 400

        value = data.get("value")
        if not value:
            return "Invalid gateway message", 400
        
        device_list = value.get("device_list")
        if not device_list:
            return "Invalid gateway message", 400
        
        for device in device_list:
            
            device_data = device.get("data")
            if not device_data:
                return "Invalid gateway message", 400
            
            scan_time = device.get("scan_time")
            if not scan_time:        
                return "Invalid gateway message", 400
            
            scan_time = str(scan_time)
            
            device_data = device_data[16:] # Skip header

            unprocessed_echobeacon_id = device_data[:4]
            packet_type = device_data[4:6]
            packet_counter = device_data[6:8]
            mac_address = device_data[8:20]
            unprocessed_rssi = device_data[20:22]

            # Only consider echo packets
            if packet_type != "03":
                continue

            echobeacon_id = str(int(unprocessed_echobeacon_id, 16))
            rssi_dbm = str(int(unprocessed_rssi, 16) - 256)

            print_formatted_data(mac_address, echobeacon_id, rssi_dbm, packet_counter, scan_time)
        
        return "Processed correctly", 200


# Used to determine remaining battery percentage

if flask_env == 'production':  
    BATTERY_MAX_MILLIVOLTS = 3600

    @connected_to_database
    def process_device(curs, device):
        # Parse raw payload
        device_data = device.get("data")
        if not device_data:
            return
        device_data = device_data[16:]  # skip gateway header
        if len(device_data) < 28:
            return
        if device_data[4:6] != "03":  # presence packets only
            return
        if not (int(device_data[22:24], 16) & 0x04):  # require Eddystone TLM flag
            return

        # Decode fields
        beacon_mac_address = device_data[8:20]
        packet_counter = int(device_data[6:8], 16)
        eddy_volt = int(device_data[24:28], 16)
        remaining_battery_percentage = round((eddy_volt / BATTERY_MAX_MILLIVOLTS) * 100, 1)
        echobeacon_friendly = int(device_data[0:4], 16)

        # Resolve tag
        curs.execute("""
            SELECT id, packet_counter, previous_echobeacon
            FROM tag
            WHERE mac_address = %s
        """, (beacon_mac_address,))
        tag = curs.fetchone()
        if not tag:
            return

        tag_id = tag['id']
        old_packet_counter = tag['packet_counter']
        prev_echobeacon_before = tag['previous_echobeacon']

        # Update telemetry; set previous_echobeacon to current beacon only when packet changes
        curs.execute("""
            UPDATE tag
            SET remaining_battery = %s,
                packet_counter = CASE
                    WHEN packet_counter IS NULL OR %s <> packet_counter THEN %s
                    ELSE packet_counter
                END,
                previous_echobeacon = CASE
                    WHEN packet_counter IS NULL OR %s <> packet_counter
                        THEN (SELECT id FROM activator_beacon WHERE friendly_number = %s)
                    ELSE previous_echobeacon
                END
            WHERE id = %s
        """, (
            remaining_battery_percentage,
            packet_counter, packet_counter,
            packet_counter, echobeacon_friendly,
            tag_id
        ))

        # Ignore duplicates and first-ever packets (need a pair)
        if old_packet_counter == packet_counter:
            return
        if old_packet_counter is None:
            return

        # Need a prior beacon to compute direction
        previous_echobeacon_id = prev_echobeacon_before
        if not previous_echobeacon_id:
            return

        # Current beacon (by friendly number)
        curs.execute("""
            SELECT id, shipyard_id, is_first_when_entering
            FROM activator_beacon
            WHERE friendly_number = %s
        """, (echobeacon_friendly,))
        current_beacon = curs.fetchone()
        if not current_beacon:
            return
        current_beacon_id = current_beacon['id']
        current_shipyard_id = current_beacon['shipyard_id']
        current_is_first_when_entering = current_beacon['is_first_when_entering']

        # No movement across same beacon
        if previous_echobeacon_id == current_beacon_id:
            return

        # Previous beacon (by stored id)
        curs.execute("""
            SELECT shipyard_id, is_first_when_entering
            FROM activator_beacon
            WHERE id = %s
        """, (previous_echobeacon_id,))
        previous_beacon = curs.fetchone()
        if not previous_beacon:
            return
        previous_shipyard_id = previous_beacon['shipyard_id']
        previous_is_first_when_entering = previous_beacon['is_first_when_entering']

        # Ignore cross-yard transitions
        if current_shipyard_id != previous_shipyard_id:
            return

        # Direction from beacon pair
        if previous_is_first_when_entering and not current_is_first_when_entering:
            is_direction_entering = True
        elif (not previous_is_first_when_entering) and current_is_first_when_entering:
            is_direction_entering = False
        else:
            return  # invalid pair (e.g., two firsts or two seconds)

        # Resolve assigned crew member (if any)
        curs.execute("""
            SELECT id
            FROM crew_member
            WHERE tag_id = %s
        """, (tag_id,))
        crew_member = curs.fetchone()

        if not crew_member:
            # Unassigned tag → record and reset pairing (prevents overlapping pairs)
            curs.execute("""
                INSERT INTO unassigned_tag_entry (tag_id, shipyard_id, is_entering)
                VALUES (%s, %s, %s)
            """, (tag_id, current_shipyard_id, is_direction_entering))
            curs.execute("UPDATE tag SET previous_echobeacon = NULL WHERE id = %s", (tag_id,))
            return

        crew_member_id = crew_member['id']

        if is_direction_entering:
            # Open new log, then reset pairing
            curs.execute("""
                INSERT INTO permanence_log (crew_member_id, shipyard_id, entry_timestamp)
                VALUES (%s, %s, NOW())
            """, (crew_member_id, current_shipyard_id))
            curs.execute("UPDATE tag SET previous_echobeacon = NULL WHERE id = %s", (tag_id,))
        else:
            # Close most recent open log (or create exit-only), then reset pairing
            curs.execute("""
                UPDATE permanence_log
                SET leave_timestamp = NOW()
                WHERE id = (
                    SELECT id FROM permanence_log
                    WHERE crew_member_id = %s
                    AND shipyard_id = %s
                    AND leave_timestamp IS NULL
                    ORDER BY entry_timestamp DESC
                    LIMIT 1
                )
            """, (crew_member_id, current_shipyard_id))
            if curs.rowcount == 0:
                curs.execute("""
                    INSERT INTO permanence_log (crew_member_id, shipyard_id, leave_timestamp)
                    VALUES (%s, %s, NOW())
                """, (crew_member_id, current_shipyard_id))
            curs.execute("UPDATE tag SET previous_echobeacon = NULL WHERE id = %s", (tag_id,))


    @app.route('/gateway-endpoint', methods=['POST'])
    def gateway_endpoint():

        json_data = request.json

        data = json_data.get("data")
        if not data:
            return "Invalid gateway message", 400

        value = data.get("value")
        if not value:
            return "Invalid gateway message", 400
        
        device_list = value.get("device_list")
        if not device_list:
            return "Invalid gateway message", 400

        for device in device_list:
            process_device(device)

        return "Processed", 200

if flask_env == "json":
    import json
    
    @app.route('/gateway-endpoint', methods=['POST'])
    def gateway_endpoint():
        json_data = request.json
        print(json.dumps(json_data, indent=2))
        return "Processed", 200


if __name__ == "__main__":
    flask_port = int(get_env("FLASK_PORT"))

    match flask_env:
        case "development":
            while True:
                try:
                    app.run(debug=True, port=flask_port, host="0.0.0.0")
                except SystemExit:
                    db_pool.close(timeout=0)
        case "production":
            serve(app, port=flask_port, host="0.0.0.0")
        case _:
            raise Exception(f"{flask_env} must be either \"development\" or \"production\"")